from openpyxl import Workbook, load_workbook

referansExcel = load_workbook("Kopya-Mevcutstoklistesi.xlsx")
referansExcelS = referansExcel.active
kayitExcel = load_workbook("hammaddeler.xlsx")
kayitExcelS = kayitExcel.active

cont = 0

for satirKont in range(2, referansExcelS.max_row):
    if (str(referansExcelS.cell(satirKont,6).value) == "MA"):
        cont = cont + 1
        i=1
        while(i<=24):
            kayitExcelS.cell(cont,i).value = referansExcelS.cell(satirKont,i).value  # type: ignore
            i = i + 1

kayitExcel.save("hammaddeler.xlsx")
            